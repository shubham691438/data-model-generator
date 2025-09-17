#!/usr/bin/env python3
"""
BRD Processing Script for Data Model Generator
Processes BRD Excel files and extracts data for solution document generation
"""

import pandas as pd
import json
import os
from pathlib import Path
from datetime import datetime
import openpyxl
from docx import Document

class BRDProcessor:
    def __init__(self, brd_directory="BRD-provided", output_directory="solution-doc-generated"):
        self.brd_directory = Path(brd_directory)
        self.output_directory = Path(output_directory)
        self.output_directory.mkdir(exist_ok=True)
        
    def process_excel_brd(self, file_path):
        """Process Excel BRD file and extract data"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            data = {}
            
            # Define expected sheet names and their processing methods
            sheet_processors = {
                'Customer Details': self._process_customer_details,
                'Features Request': self._process_features_request,
                'Pull Questions from the ATS': self._process_ats_questions,
                'Apply Form': self._process_apply_form,
                'Set Questions Manually': self._process_manual_questions,
                'Source Tracking': self._process_source_tracking,
                'Domain Access': self._process_domain_access,
                'Static Client Level Question': self._process_static_questions,
                'Job Filter Level Questions': self._process_job_filter_questions
            }
            
            # Process each sheet
            for sheet_name, processor in sheet_processors.items():
                if sheet_name in workbook.sheetnames:
                    try:
                        sheet_data = processor(workbook[sheet_name])
                        data[sheet_name] = sheet_data
                    except Exception as e:
                        print(f"Warning: Error processing sheet '{sheet_name}': {e}")
                        data[sheet_name] = {}
                else:
                    print(f"Warning: Sheet '{sheet_name}' not found in {file_path}")
                    data[sheet_name] = {}
            
            return data
            
        except Exception as e:
            print(f"Error processing Excel file {file_path}: {e}")
            return {}
    
    def process_docx_brd(self, file_path):
        """Process Word BRD file and extract data"""
        try:
            doc = Document(file_path)
            data = {
                'Customer Details': {},
                'Features Request': {},
                'Integration Type': 'S2S Funnel Tracking',
                'Document Type': 'Word Document',
                'Raw Content': []
            }
            
            # Extract text content
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    data['Raw Content'].append(paragraph.text.strip())
            
            # Try to extract structured information
            content = ' '.join(data['Raw Content'])
            
            # Extract client name from filename or content
            filename = Path(file_path).stem
            if 'Charter' in filename or 'Spectrum' in filename:
                data['Customer Details']['Client Name'] = 'Charter Spectrum'
            
            # Look for integration type
            if 'S2S' in content or 'Server-to-Server' in content:
                data['Integration Type'] = 'S2S Funnel Tracking'
            
            return data
            
        except Exception as e:
            print(f"Error processing Word file {file_path}: {e}")
            return {}
    
    def _process_customer_details(self, sheet):
        """Process Customer Details sheet"""
        data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                data[row[0]] = row[1]
        return data
    
    def _process_features_request(self, sheet):
        """Process Features Request sheet"""
        features = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and len(row) > 1:
                feature_name = row[0]
                status = row[1] if row[1] else 'No'
                features[feature_name] = status
        return features
    
    def _process_ats_questions(self, sheet):
        """Process ATS Questions sheet"""
        questions = []
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                question_data = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i] and value:
                        question_data[headers[i]] = value
                if question_data:
                    questions.append(question_data)
        
        return questions
    
    def _process_apply_form(self, sheet):
        """Process Apply Form sheet"""
        form_config = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                form_config[row[0]] = row[1]
        return form_config
    
    def _process_manual_questions(self, sheet):
        """Process Manual Questions sheet"""
        questions = []
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                question_data = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i] and value:
                        question_data[headers[i]] = value
                if question_data:
                    questions.append(question_data)
        
        return questions
    
    def _process_source_tracking(self, sheet):
        """Process Source Tracking sheet"""
        tracking_config = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                tracking_config[row[0]] = row[1]
        return tracking_config
    
    def _process_domain_access(self, sheet):
        """Process Domain Access sheet"""
        domain_config = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                domain_config[row[0]] = row[1]
        return domain_config
    
    def _process_static_questions(self, sheet):
        """Process Static Client Level Questions sheet"""
        questions = []
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                question_data = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i] and value:
                        question_data[headers[i]] = value
                if question_data:
                    questions.append(question_data)
        
        return questions
    
    def _process_job_filter_questions(self, sheet):
        """Process Job Filter Level Questions sheet"""
        questions = []
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                question_data = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i] and value:
                        question_data[headers[i]] = value
                if question_data:
                    questions.append(question_data)
        
        return questions
    
    def process_all_brds(self):
        """Process all BRD files in the directory"""
        results = {}
        
        # Process Excel files
        for file_path in self.brd_directory.glob("*.xlsx"):
            print(f"Processing Excel BRD: {file_path}")
            data = self.process_excel_brd(file_path)
            if data:
                results[file_path.stem] = data
        
        # Process Word files
        for file_path in self.brd_directory.glob("*.docx"):
            print(f"Processing Word BRD: {file_path}")
            data = self.process_docx_brd(file_path)
            if data:
                results[file_path.stem] = data
        
        # Save processed data
        output_file = self.output_directory / "processed_brd_data.json"
        with open(output_file, 'w') as f:
            json.dump(results, f, indent=2, default=str)
        
        print(f"Processed data saved to: {output_file}")
        return results

if __name__ == "__main__":
    processor = BRDProcessor()
    results = processor.process_all_brds()
    print(f"Processed {len(results)} BRD files")
